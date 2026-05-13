import csv
from io import BytesIO, StringIO
import re

from fastapi import Depends, FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from sqlalchemy import text
from sqlalchemy.orm import Session, joinedload
import random
from datetime import datetime, timezone

from .audit import write_audit
from .version import __version__
from .auth import (
    authenticate_local_user,
    build_totp_provisioning_uri,
    create_access_token,
    ensure_bootstrap_admin,
    generate_totp_secret,
    get_current_identity,
    get_current_user,
    ldap_authenticate,
    normalize_otp_code,
    require_admin,
    hash_password,
    verify_totp_code,
)
from .config import settings
from .database import Base, engine, get_db
from .models import (
    AuditLog,
    Device,
    DeviceModel,
    Floorplan,
    InventoryDevice,
    LocalUser,
    Rack,
    ServerRoom,
    ServerRoomFloorplan,
)
from .schemas import (
    AuditOut,
    DeviceCreate,
    DeviceModelCreate,
    DeviceModelOut,
    DeviceOut,
    DeviceUpdate,
    FloorplanCreate,
    FloorplanOut,
    FloorplanUpdate,
    LocalUserCreate,
    LocalUserOut,
    LoginRequest,
    PasswordConfirmRequest,
    InventoryDeviceOut,
    RackCreate,
    RackOut,
    RackUpdate,
    ServerRoomCreate,
    ServerRoomOut,
    TwoFactorCodeRequest,
    TwoFactorSetupOut,
    TwoFactorStatusOut,
    TokenResponse,
)

Base.metadata.create_all(bind=engine)


def ensure_mount_side_columns() -> None:
    with engine.begin() as conn:
        conn.execute(
            text(
                "ALTER TABLE devices ADD COLUMN IF NOT EXISTS mount_side VARCHAR(16) NOT NULL DEFAULT 'front'"
            )
        )
        conn.execute(
            text(
                "ALTER TABLE inventory_devices ADD COLUMN IF NOT EXISTS mount_side VARCHAR(16) NOT NULL DEFAULT 'front'"
            )
        )


def ensure_inventory_archive_columns() -> None:
    with engine.begin() as conn:
        conn.execute(
            text(
                "ALTER TABLE inventory_devices ADD COLUMN IF NOT EXISTS archived INTEGER NOT NULL DEFAULT 0"
            )
        )
        conn.execute(
            text(
                "ALTER TABLE inventory_devices ADD COLUMN IF NOT EXISTS archived_at TIMESTAMPTZ NULL"
            )
        )


def ensure_local_user_totp_columns() -> None:
    with engine.begin() as conn:
        conn.execute(
            text(
                "ALTER TABLE local_users ADD COLUMN IF NOT EXISTS totp_secret VARCHAR(64) NULL"
            )
        )
        conn.execute(
            text(
                "ALTER TABLE local_users ADD COLUMN IF NOT EXISTS totp_enabled INTEGER NOT NULL DEFAULT 0"
            )
        )


ensure_mount_side_columns()
ensure_inventory_archive_columns()
ensure_local_user_totp_columns()


SHEET_NAME_MAX_LENGTH = 31


def sanitize_excel_sheet_name(name: str, fallback: str) -> str:
    sanitized = re.sub(r"[\\/*?:\[\]]", "_", (name or "").strip())
    sanitized = sanitized.strip("'")
    if not sanitized:
        sanitized = fallback
    return sanitized[:SHEET_NAME_MAX_LENGTH]


def unique_sheet_name(name: str, seen: set[str], fallback: str) -> str:
    base_name = sanitize_excel_sheet_name(name, fallback)
    candidate = base_name
    suffix = 1
    while candidate in seen:
        suffix_text = f"_{suffix}"
        candidate = f"{base_name[:SHEET_NAME_MAX_LENGTH - len(suffix_text)]}{suffix_text}"
        suffix += 1
    seen.add(candidate)
    return candidate


def build_rack_u_rows(rack: Rack) -> list[dict[str, str | int]]:
    units = rack.units or 42
    device_by_u: dict[int, Device] = {}
    for device in rack.devices:
        start_u = max(1, device.u_position)
        end_u = min(units, device.u_position + max(1, device.u_height) - 1)
        for u_value in range(start_u, end_u + 1):
            device_by_u[u_value] = device

    rows: list[dict[str, str | int]] = []
    for u_value in range(units, 0, -1):
        device = device_by_u.get(u_value)
        properties = device.properties if device else {}
        rows.append(
            {
                "u": u_value,
                "device_name": device.name if device else "",
                "device_type": device.device_type if device else "",
                "model": (device.model or "") if device else "",
                "vendor": (device.vendor or "") if device else "",
                "serial_number": (device.serial_number or "") if device else "",
                "management_ip": (device.management_ip or "") if device else "",
                "mount_side": (device.mount_side or "front") if device else "",
                "u_height": device.u_height if device else "",
                "device_start_u": device.u_position if device else "",
                "hostname": str(properties.get("hostname", "")) if properties else "",
                "host_ip": str(properties.get("host_ip", "")) if properties else "",
                "ssh_endpoint": str(properties.get("ssh_endpoint", "")) if properties else "",
                "notes": str(properties.get("notes", "")) if properties else "",
            }
        )
    return rows


def apply_rack_device_merges(sheet, rack: Rack, data_start_row: int, end_column: int) -> None:
    units = rack.units or 42
    for device in rack.devices:
        start_u = max(1, device.u_position)
        end_u = min(units, device.u_position + max(1, device.u_height) - 1)
        span = end_u - start_u + 1
        if span <= 1:
            continue

        top_row = data_start_row + (units - end_u)
        bottom_row = data_start_row + (units - start_u)
        for column in range(2, end_column + 1):
            sheet.merge_cells(start_row=top_row, start_column=column, end_row=bottom_row, end_column=column)
            sheet.cell(row=top_row, column=column).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def build_serverroom_export_workbook(room: ServerRoom) -> BytesIO:
    workbook = Workbook()
    workbook.remove(workbook.active)
    used_sheet_names: set[str] = set()

    floorplans = sorted(
        [link.floorplan for link in room.floorplan_links if link.floorplan],
        key=lambda floorplan: floorplan.id,
    )
    racks = sorted(
        [rack for floorplan in floorplans for rack in floorplan.racks],
        key=lambda rack: (rack.name.lower(), rack.id),
    )

    if not racks:
        sheet = workbook.create_sheet(title="Rack_Export")
        sheet.append(["Serverroom", room.name])
        sheet.append(["Message", "No racks found for this serverroom"])
    else:
        header = [
            "U",
            "Device Name",
            "Device Type",
            "Model",
            "Vendor",
            "Serial Number",
            "Management IP",
            "Mount Side",
            "Device Height U",
            "Device Start U",
            "Hostname",
            "Host IP",
            "SSH Endpoint",
            "Notes",
        ]

        for rack in racks:
            sheet_name = unique_sheet_name(rack.name, used_sheet_names, f"Rack_{rack.id}")
            sheet = workbook.create_sheet(title=sheet_name)
            floorplan_name = rack.floorplan.name if rack.floorplan else ""
            sheet.append(["Serverroom", room.name])
            sheet.append(["Floorplan", floorplan_name])
            sheet.append(["Rack", rack.name])
            sheet.append(["Units", rack.units])
            sheet.append([])
            sheet.append(header)

            for row in build_rack_u_rows(rack):
                sheet.append(
                    [
                        row["u"],
                        row["device_name"],
                        row["device_type"],
                        row["model"],
                        row["vendor"],
                        row["serial_number"],
                        row["management_ip"],
                        row["mount_side"],
                        row["u_height"],
                        row["device_start_u"],
                        row["hostname"],
                        row["host_ip"],
                        row["ssh_endpoint"],
                        row["notes"],
                    ]
                )

            data_start_row = 7
            apply_rack_device_merges(sheet, rack, data_start_row=data_start_row, end_column=len(header))

            for cell in sheet[6]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            for row_index in range(data_start_row, data_start_row + (rack.units or 42)):
                sheet.cell(row=row_index, column=1).alignment = Alignment(horizontal="center", vertical="center")

            column_widths = {
                "A": 8,
                "B": 28,
                "C": 14,
                "D": 20,
                "E": 18,
                "F": 18,
                "G": 18,
                "H": 12,
                "I": 14,
                "J": 14,
                "K": 22,
                "L": 18,
                "M": 28,
                "N": 36,
            }
            for column_name, width in column_widths.items():
                sheet.column_dimensions[column_name].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def export_filename_for_serverroom(room_name: str) -> str:
    slug = re.sub(r"[^A-Za-z0-9._-]+", "-", room_name.strip()).strip("-._")
    return f"{slug or 'serverroom'}.xlsx"


def build_demo_serial(rng: random.Random) -> str:
    return f"SN{rng.randint(100000, 999999)}{rng.choice(['A', 'B', 'C', 'D'])}"


def build_demo_ip(rng: random.Random, subnet: int) -> str:
    return f"10.{subnet}.{rng.randint(1, 254)}.{rng.randint(1, 254)}"


def seed_demo_devices_for_rack(db: Session, rack: Rack, device_models: list[DeviceModel], seed: int) -> None:
    if rack.devices:
        return

    rng = random.Random(seed)
    cursor_u = 1
    item_index = 1
    target_count = rng.randint(8, 12)

    while cursor_u <= rack.units and item_index <= target_count:
        model = rng.choice(device_models)
        u_height = min(model.u_height, 3)
        if cursor_u + u_height - 1 > rack.units:
            break

        # Keep most racks front-only, but seed a few racks with front/back mix.
        if rack.name in {"A01", "A02", "B01", "B02"}:
            if model.device_type == "switch":
                mount_side = "back"
            else:
                mount_side = "back" if rng.random() < 0.35 else "front"
        else:
            mount_side = "front"

        hostname = f"{rack.name.lower().replace(' ', '-')}-n{item_index}"
        mgmt_ip = build_demo_ip(rng, 120)
        host_ip = build_demo_ip(rng, 121)
        serial = build_demo_serial(rng)
        ssh_endpoint = f"ssh://{hostname}.lab.local"

        device = Device(
            rack_id=rack.id,
            name=f"{model.name} #{item_index}",
            device_type=model.device_type,
            u_position=cursor_u,
            u_height=u_height,
            mount_side=mount_side,
            serial_number=serial,
            management_ip=mgmt_ip,
            model=model.model_code,
            vendor=model.vendor,
            properties={
                "image_url": model.image_url,
                "hostname": hostname,
                "host_ip": host_ip,
                "ssh_endpoint": ssh_endpoint,
                "notes": "Auto-seeded demo device",
            },
        )
        db.add(device)
        cursor_u += u_height + (1 if rng.random() > 0.55 else 0)
        item_index += 1


def ensure_demo_layout_for_serverroom(db: Session, room: ServerRoom) -> None:
    room_link = (
        db.query(ServerRoomFloorplan)
        .options(joinedload(ServerRoomFloorplan.floorplan))
        .filter(ServerRoomFloorplan.serverroom_id == room.id)
        .first()
    )

    if room_link is None:
        floorplan = Floorplan(name=f"{room.name} Floor", width=1000, height=640)
        db.add(floorplan)
        db.flush()
        db.add(ServerRoomFloorplan(serverroom_id=room.id, floorplan_id=floorplan.id))
    else:
        floorplan = room_link.floorplan

    rack_names = [f"A{i:02d}" for i in range(1, 9)] + [f"B{i:02d}" for i in range(1, 9)]
    existing_racks = db.query(Rack).filter(Rack.floorplan_id == floorplan.id).order_by(Rack.id.asc()).all()
    existing_by_name = {rack.name: rack for rack in existing_racks}

    start_x = 120
    start_y = 180
    spacing_x = 90
    spacing_y = 110
    width = 52
    height = 34

    for idx, rack_name in enumerate(rack_names):
        row = 0 if idx < 8 else 1
        col = idx if idx < 8 else idx - 8
        x = start_x + col * spacing_x
        y = start_y + row * spacing_y

        rack = existing_by_name.get(rack_name)
        if rack is None:
            rack = Rack(
                floorplan_id=floorplan.id,
                name=rack_name,
                x=x,
                y=y,
                width=width,
                height=height,
                units=42,
                orientation="top",
            )
            db.add(rack)
        else:
            rack.x = x
            rack.y = y
            rack.width = width
            rack.height = height
            rack.orientation = "top"

    # Normalize older front-view racks so floorplan remains readable in top-view mode.
    for rack in existing_racks:
        if rack.name in rack_names:
            continue
        if rack.height > 60 or rack.width > 90:
            rack.width = width
            rack.height = height
            rack.orientation = "top"

    db.flush()

    models = db.query(DeviceModel).all()
    if models:
        racks = db.query(Rack).filter(Rack.floorplan_id == floorplan.id).order_by(Rack.id.asc()).all()
        for rack in racks:
            seed_demo_devices_for_rack(db, rack, models, seed=(room.id * 1000) + rack.id)


def ensure_demo_layout_all_serverrooms(db: Session) -> None:
    rooms = db.query(ServerRoom).order_by(ServerRoom.id.asc()).all()
    for room in rooms:
        ensure_demo_layout_for_serverroom(db, room)

    db.commit()


def ensure_demo_mount_side_mix(db: Session) -> None:
    target_racks = {"A01", "A02", "B01", "B02"}
    racks = db.query(Rack).filter(Rack.name.in_(target_racks)).all()
    for rack in racks:
        devices = db.query(Device).filter(Device.rack_id == rack.id).order_by(Device.id.asc()).all()
        if len(devices) < 2:
            continue
        has_front = any((d.mount_side or "front") == "front" for d in devices)
        has_back = any((d.mount_side or "front") == "back" for d in devices)
        if has_front and has_back:
            continue

        rng = random.Random((rack.id * 131) + 17)
        switch_ids = [d.id for d in devices if d.device_type == "switch"]
        candidate_ids = switch_ids if switch_ids else [d.id for d in devices]
        take = max(1, min(len(candidate_ids), len(devices) // 3))
        selected = set(rng.sample(candidate_ids, take))

        for d in devices:
            d.mount_side = "back" if d.id in selected else "front"

    db.commit()


def ensure_builtin_device_models(db: Session) -> None:
    builtin_models = [
        {
            "name": "Dell PowerEdge R670",
            "vendor": "Dell",
            "model_code": "PowerEdge R670",
            "device_type": "server",
            "u_height": 1,
            "image_url": "/static/device-models/dell-r670.svg",
        },
        {
            "name": "HP ProLiant DL380",
            "vendor": "HPE",
            "model_code": "ProLiant DL380",
            "device_type": "server",
            "u_height": 2,
            "image_url": "/static/device-models/hp-dl380.svg",
        },
        {
            "name": "Cisco Catalyst 9300",
            "vendor": "Cisco",
            "model_code": "Catalyst 9300",
            "device_type": "switch",
            "u_height": 1,
            "image_url": "/static/device-models/cisco-c9300.svg",
        },
        {
            "name": "Brocade ICX 7450",
            "vendor": "Brocade",
            "model_code": "ICX 7450",
            "device_type": "switch",
            "u_height": 1,
            "image_url": "/static/device-models/brocade-icx7450.svg",
        },
    ]

    existing = {
        (m.vendor.lower(), m.model_code.lower())
        for m in db.query(DeviceModel).all()
    }

    for model in builtin_models:
        key = (model["vendor"].lower(), model["model_code"].lower())
        if key in existing:
            continue
        db.add(DeviceModel(**model))

    db.commit()

with Session(bind=engine) as bootstrap_db:
    ensure_bootstrap_admin(bootstrap_db)
    if bootstrap_db.query(ServerRoom).count() == 0:
        bootstrap_db.add(ServerRoom(name="Main Serverroom", description="Default serverroom"))
        bootstrap_db.commit()
    ensure_builtin_device_models(bootstrap_db)
    ensure_demo_layout_all_serverrooms(bootstrap_db)
    ensure_demo_mount_side_mix(bootstrap_db)

app = FastAPI(title=settings.app_name)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

app.mount("/static", StaticFiles(directory="app/static"), name="static")


@app.get("/")
def root() -> FileResponse:
    return FileResponse("app/static/index.html")


@app.get("/api/version")
def get_version():
    return {"version": __version__}


@app.post("/api/auth/login", response_model=TokenResponse)
def login(payload: LoginRequest, db: Session = Depends(get_db)) -> TokenResponse:
    local_user = authenticate_local_user(db, payload.username, payload.password)
    if local_user:
        if local_user.totp_enabled == 1 and not verify_totp_code(local_user.totp_secret, payload.otp_code):
            raise HTTPException(status_code=401, detail="2FA code required or invalid")

        token, expires_at = create_access_token(local_user.username, role=local_user.role, auth_source="local")
        return TokenResponse(
            access_token=token,
            username=local_user.username,
            role=local_user.role,
            auth_source="local",
            expires_at=expires_at,
        )

    if ldap_authenticate(payload.username, payload.password):
        token, expires_at = create_access_token(payload.username, role="user", auth_source="ldap")
        return TokenResponse(
            access_token=token,
            username=payload.username,
            role="user",
            auth_source="ldap",
            expires_at=expires_at,
        )

    raise HTTPException(status_code=401, detail="Invalid credentials")


@app.get("/api/auth/2fa/status", response_model=TwoFactorStatusOut)
def get_two_factor_status(
    db: Session = Depends(get_db),
    identity: dict[str, str] = Depends(get_current_identity),
) -> TwoFactorStatusOut:
    if identity.get("auth_source") != "local":
        return TwoFactorStatusOut(available=False, enabled=False, setup_pending=False)

    user = db.query(LocalUser).filter(LocalUser.username == identity["sub"]).first()
    if not user:
        raise HTTPException(status_code=404, detail="Local user not found")

    return TwoFactorStatusOut(
        available=True,
        enabled=user.totp_enabled == 1,
        setup_pending=bool(user.totp_secret) and user.totp_enabled != 1,
    )


@app.post("/api/auth/2fa/setup", response_model=TwoFactorSetupOut)
def setup_two_factor(
    db: Session = Depends(get_db),
    identity: dict[str, str] = Depends(get_current_identity),
) -> TwoFactorSetupOut:
    if identity.get("auth_source") != "local":
        raise HTTPException(status_code=400, detail="2FA setup is available only for local accounts")

    user = db.query(LocalUser).filter(LocalUser.username == identity["sub"]).first()
    if not user:
        raise HTTPException(status_code=404, detail="Local user not found")

    secret = generate_totp_secret()
    user.totp_secret = secret
    user.totp_enabled = 0
    db.commit()
    return TwoFactorSetupOut(secret=secret, provisioning_uri=build_totp_provisioning_uri(secret, user.username))


@app.post("/api/auth/2fa/confirm", response_model=TwoFactorStatusOut)
def confirm_two_factor(
    payload: TwoFactorCodeRequest,
    db: Session = Depends(get_db),
    identity: dict[str, str] = Depends(get_current_identity),
) -> TwoFactorStatusOut:
    if identity.get("auth_source") != "local":
        raise HTTPException(status_code=400, detail="2FA setup is available only for local accounts")

    user = db.query(LocalUser).filter(LocalUser.username == identity["sub"]).first()
    if not user:
        raise HTTPException(status_code=404, detail="Local user not found")
    if not user.totp_secret:
        raise HTTPException(status_code=400, detail="Start 2FA setup first")
    if not verify_totp_code(user.totp_secret, normalize_otp_code(payload.otp_code)):
        raise HTTPException(status_code=400, detail="Invalid 2FA code")

    user.totp_enabled = 1
    write_audit(
        db,
        actor=identity["sub"],
        action="enable_2fa",
        entity_type="local_user",
        entity_id=str(user.id),
        new_values={"totp_enabled": 1},
    )
    db.commit()
    return TwoFactorStatusOut(available=True, enabled=True, setup_pending=False)


@app.post("/api/auth/2fa/disable", response_model=TwoFactorStatusOut)
def disable_two_factor(
    payload: TwoFactorCodeRequest,
    db: Session = Depends(get_db),
    identity: dict[str, str] = Depends(get_current_identity),
) -> TwoFactorStatusOut:
    if identity.get("auth_source") != "local":
        raise HTTPException(status_code=400, detail="2FA setup is available only for local accounts")

    user = db.query(LocalUser).filter(LocalUser.username == identity["sub"]).first()
    if not user:
        raise HTTPException(status_code=404, detail="Local user not found")
    if user.totp_enabled != 1 or not user.totp_secret:
        raise HTTPException(status_code=400, detail="2FA is not enabled")
    if not verify_totp_code(user.totp_secret, normalize_otp_code(payload.otp_code)):
        raise HTTPException(status_code=400, detail="Invalid 2FA code")

    user.totp_secret = None
    user.totp_enabled = 0
    write_audit(
        db,
        actor=identity["sub"],
        action="disable_2fa",
        entity_type="local_user",
        entity_id=str(user.id),
        new_values={"totp_enabled": 0},
    )
    db.commit()
    return TwoFactorStatusOut(available=True, enabled=False, setup_pending=False)


def map_floorplan_serverroom_id(floorplan: Floorplan) -> int | None:
    if not floorplan.serverroom_links:
        return None
    return floorplan.serverroom_links[0].serverroom_id


def normalize_csv_key(value: str) -> str:
    return "".join(ch for ch in value.strip().lower() if ch.isalnum())


def csv_value(row: dict[str, str], aliases: list[str], default: str = "") -> str:
    normalized = {normalize_csv_key(k): (v or "").strip() for k, v in row.items()}
    for alias in aliases:
        value = normalized.get(normalize_csv_key(alias), "")
        if value != "":
            return value
    return default


def csv_int_value(row: dict[str, str], aliases: list[str], default: int = 0) -> int:
    raw = csv_value(row, aliases)
    if raw == "":
        return default
    try:
        return int(raw)
    except ValueError as err:
        raise ValueError(f"Invalid integer value '{raw}'") from err


def parse_csv_upload(file: UploadFile) -> list[dict[str, str]]:
    content = file.file.read()
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as err:
        raise HTTPException(status_code=400, detail="CSV must be UTF-8 encoded") from err

    reader = csv.DictReader(StringIO(text))
    if not reader.fieldnames:
        raise HTTPException(status_code=400, detail="CSV file is missing header row")
    return list(reader)


def find_serverroom_by_name(db: Session, name: str) -> ServerRoom | None:
    needle = name.strip().lower()
    if needle == "":
        return None
    return db.query(ServerRoom).filter(ServerRoom.name.ilike(name.strip())).first()


def find_floorplan_by_room_and_name(db: Session, room_id: int, floor_name: str) -> Floorplan | None:
    return (
        db.query(Floorplan)
        .join(ServerRoomFloorplan, ServerRoomFloorplan.floorplan_id == Floorplan.id)
        .filter(ServerRoomFloorplan.serverroom_id == room_id)
        .filter(Floorplan.name.ilike(floor_name.strip()))
        .first()
    )


def find_rack_by_floorplan_and_name(db: Session, floorplan_id: int, rack_name: str) -> Rack | None:
    return (
        db.query(Rack)
        .filter(Rack.floorplan_id == floorplan_id)
        .filter(Rack.name.ilike(rack_name.strip()))
        .first()
    )


def validate_device_placement(
    db: Session,
    rack_id: int,
    u_position: int,
    u_height: int,
    mount_side: str,
    exclude_device_id: int | None = None,
) -> None:
    rack = db.query(Rack).filter(Rack.id == rack_id).first()
    if not rack:
        raise HTTPException(status_code=404, detail="Rack not found")

    if u_position < 1 or u_height < 1:
        raise HTTPException(status_code=400, detail="U position and U height must be >= 1")

    if mount_side not in {"front", "back"}:
        raise HTTPException(status_code=400, detail="mount_side must be 'front' or 'back'")

    if (u_position + u_height - 1) > rack.units:
        raise HTTPException(status_code=400, detail="Device exceeds rack unit capacity")

    devices_query = db.query(Device).filter(Device.rack_id == rack_id)
    if exclude_device_id is not None:
        devices_query = devices_query.filter(Device.id != exclude_device_id)
    existing_devices = devices_query.all()

    new_start = u_position
    new_end = u_position + u_height - 1
    for other in existing_devices:
        if (other.mount_side or "front") != mount_side:
            continue
        other_start = other.u_position
        other_end = other.u_position + other.u_height - 1
        if max(new_start, other_start) <= min(new_end, other_end):
            raise HTTPException(status_code=409, detail=f"U overlap with device '{other.name}'")


def upsert_inventory_from_device(db: Session, device: Device, archived: bool = False) -> str:
    serial = (device.serial_number or "").strip() or f"UNRACKED-{device.id}"
    properties = device.properties or {}
    archived_flag = 1 if archived else 0
    archived_at_value = datetime.now(timezone.utc) if archived else None

    existing = db.query(InventoryDevice).filter(InventoryDevice.serial_number == serial).first()
    if existing:
        existing.name = device.name
        existing.model = device.model
        existing.vendor = device.vendor
        existing.device_type = device.device_type
        existing.u_height = device.u_height
        existing.mount_side = device.mount_side or "front"
        existing.management_ip = device.management_ip
        existing.archived = archived_flag
        existing.archived_at = archived_at_value
        existing.properties = properties
    else:
        db.add(
            InventoryDevice(
                serial_number=serial,
                name=device.name,
                model=device.model,
                vendor=device.vendor,
                device_type=device.device_type,
                u_height=device.u_height,
                mount_side=device.mount_side or "front",
                management_ip=device.management_ip,
                archived=archived_flag,
                archived_at=archived_at_value,
                properties=properties,
            )
        )
    return serial


def ensure_serial_not_archived(db: Session, serial_number: str | None) -> None:
    serial = (serial_number or "").strip()
    if not serial:
        return
    archived_inventory = (
        db.query(InventoryDevice)
        .filter(InventoryDevice.serial_number == serial)
        .filter(InventoryDevice.archived == 1)
        .first()
    )
    if archived_inventory:
        raise HTTPException(status_code=409, detail=f"Device serial '{serial}' is archived and cannot be assigned")


@app.get("/api/serverrooms", response_model=list[ServerRoomOut])
def list_serverrooms(
    db: Session = Depends(get_db),
    _: str = Depends(get_current_user),
) -> list[ServerRoomOut]:
    return db.query(ServerRoom).order_by(ServerRoom.name.asc()).all()


@app.get("/api/serverrooms/{serverroom_id}/export.xlsx")
def export_serverroom_xlsx(
    serverroom_id: int,
    db: Session = Depends(get_db),
    _: str = Depends(get_current_user),
) -> StreamingResponse:
    room = (
        db.query(ServerRoom)
        .options(
            joinedload(ServerRoom.floorplan_links)
            .joinedload(ServerRoomFloorplan.floorplan)
            .joinedload(Floorplan.racks)
            .joinedload(Rack.devices),
            joinedload(ServerRoom.floorplan_links)
            .joinedload(ServerRoomFloorplan.floorplan)
            .joinedload(Floorplan.racks)
            .joinedload(Rack.floorplan),
        )
        .filter(ServerRoom.id == serverroom_id)
        .first()
    )
    if not room:
        raise HTTPException(status_code=404, detail="Serverroom not found")

    workbook = build_serverroom_export_workbook(room)
    filename = export_filename_for_serverroom(room.name)
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        workbook,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@app.post("/api/serverrooms", response_model=ServerRoomOut)
def create_serverroom(
    payload: ServerRoomCreate,
    db: Session = Depends(get_db),
    user: str = Depends(get_current_user),
) -> ServerRoomOut:
    existing = db.query(ServerRoom).filter(ServerRoom.name == payload.name).first()
    if existing:
        raise HTTPException(status_code=409, detail="Serverroom name already exists")

    room = ServerRoom(**payload.model_dump())
    db.add(room)
    db.flush()
    write_audit(
        db,
        actor=user,
        action="create",
        entity_type="serverroom",
        entity_id=str(room.id),
        new_values=payload.model_dump(),
    )
    db.commit()
    db.refresh(room)
    ensure_demo_layout_for_serverroom(db, room)
    db.commit()
    return room


@app.delete("/api/serverrooms/{serverroom_id}")
def delete_serverroom(
    serverroom_id: int,
    db: Session = Depends(get_db),
    user: str = Depends(get_current_user),
) -> dict[str, bool]:
    room = db.query(ServerRoom).filter(ServerRoom.id == serverroom_id).first()
    if not room:
        raise HTTPException(status_code=404, detail="Serverroom not found")

    old_values = {
        "name": room.name,
        "description": room.description,
    }
    db.delete(room)
    write_audit(
        db,
        actor=user,
        action="delete",
        entity_type="serverroom",
        entity_id=str(serverroom_id),
        old_values=old_values,
    )
    db.commit()
    return {"ok": True}


@app.get("/api/device-models", response_model=list[DeviceModelOut])
def list_device_models(
    db: Session = Depends(get_db),
    _: str = Depends(get_current_user),
) -> list[DeviceModelOut]:
    return db.query(DeviceModel).order_by(DeviceModel.vendor.asc(), DeviceModel.name.asc()).all()


@app.get("/api/inventory-devices", response_model=list[InventoryDeviceOut])
def list_inventory_devices(
    include_archived: bool = True,
    db: Session = Depends(get_db),
    _: str = Depends(get_current_user),
) -> list[InventoryDeviceOut]:
    query = db.query(InventoryDevice)
    if not include_archived:
        query = query.filter(InventoryDevice.archived == 0)
    return query.order_by(InventoryDevice.updated_at.desc(), InventoryDevice.id.desc()).all()


@app.post("/api/device-models", response_model=DeviceModelOut)
def create_device_model(
    payload: DeviceModelCreate,
    db: Session = Depends(get_db),
    user: str = Depends(get_current_user),
) -> DeviceModelOut:
    model = DeviceModel(**payload.model_dump())
    db.add(model)
    db.flush()
    write_audit(
        db,
        actor=user,
        action="create",
        entity_type="device_model",
        entity_id=str(model.id),
        new_values=payload.model_dump(),
    )
    db.commit()
    db.refresh(model)
    return model


@app.get("/api/local-users", response_model=list[LocalUserOut])
def list_local_users(
    db: Session = Depends(get_db),
    _: str = Depends(require_admin),
) -> list[LocalUserOut]:
    users = db.query(LocalUser).order_by(LocalUser.username.asc()).all()
    return users


@app.post("/api/local-users", response_model=LocalUserOut)
def create_local_user(
    payload: LocalUserCreate,
    db: Session = Depends(get_db),
    actor: str = Depends(require_admin),
) -> LocalUserOut:
    existing = db.query(LocalUser).filter(LocalUser.username == payload.username).first()
    if existing:
        raise HTTPException(status_code=409, detail="Username already exists")

    if payload.role not in {"admin", "user"}:
        raise HTTPException(status_code=400, detail="Role must be 'admin' or 'user'")

    user = LocalUser(
        username=payload.username,
        password_hash=hash_password(payload.password),
        role=payload.role,
        is_active=1,
    )
    db.add(user)
    db.flush()
    write_audit(
        db,
        actor=actor,
        action="create",
        entity_type="local_user",
        entity_id=str(user.id),
        new_values={"username": user.username, "role": user.role, "is_active": user.is_active},
    )
    db.commit()
    db.refresh(user)
    return user


@app.get("/api/floorplans", response_model=list[FloorplanOut])
def list_floorplans(
    serverroom_id: int | None = None,
    db: Session = Depends(get_db),
    _: str = Depends(get_current_user),
) -> list[FloorplanOut]:
    query = (
        db.query(Floorplan)
        .options(
            joinedload(Floorplan.racks).joinedload(Rack.devices),
            joinedload(Floorplan.serverroom_links),
        )
        .order_by(Floorplan.id.asc())
    )

    if serverroom_id is not None:
        query = query.join(ServerRoomFloorplan).filter(ServerRoomFloorplan.serverroom_id == serverroom_id)

    floorplans = query.all()
    result: list[FloorplanOut] = []
    for floorplan in floorplans:
        out = FloorplanOut.model_validate(floorplan)
        out.serverroom_id = map_floorplan_serverroom_id(floorplan)
        result.append(out)
    return result


@app.post("/api/floorplans", response_model=FloorplanOut)
def create_floorplan(
    payload: FloorplanCreate,
    db: Session = Depends(get_db),
    user: str = Depends(get_current_user),
) -> FloorplanOut:
    room = db.query(ServerRoom).filter(ServerRoom.id == payload.serverroom_id).first()
    if not room:
        raise HTTPException(status_code=404, detail="Serverroom not found")

    floorplan = Floorplan(name=payload.name, width=payload.width, height=payload.height)
    db.add(floorplan)
    db.flush()
    db.add(ServerRoomFloorplan(serverroom_id=payload.serverroom_id, floorplan_id=floorplan.id))
    write_audit(
        db,
        actor=user,
        action="create",
        entity_type="floorplan",
        entity_id=str(floorplan.id),
        new_values=payload.model_dump(),
    )
    db.commit()
    db.refresh(floorplan)
    out = FloorplanOut.model_validate(floorplan)
    out.serverroom_id = payload.serverroom_id
    return out


@app.delete("/api/floorplans/{floorplan_id}")
def delete_floorplan(
    floorplan_id: int,
    db: Session = Depends(get_db),
    user: str = Depends(get_current_user),
) -> dict[str, bool]:
    floorplan = db.query(Floorplan).filter(Floorplan.id == floorplan_id).first()
    if not floorplan:
        raise HTTPException(status_code=404, detail="Floorplan not found")

    old_values = {
        "name": floorplan.name,
        "width": floorplan.width,
        "height": floorplan.height,
    }
    db.delete(floorplan)
    write_audit(
        db,
        actor=user,
        action="delete",
        entity_type="floorplan",
        entity_id=str(floorplan_id),
        old_values=old_values,
    )
    db.commit()
    return {"ok": True}


@app.put("/api/floorplans/{floorplan_id}", response_model=FloorplanOut)
def update_floorplan(
    floorplan_id: int,
    payload: FloorplanUpdate,
    db: Session = Depends(get_db),
    user: str = Depends(get_current_user),
) -> FloorplanOut:
    floorplan = (
        db.query(Floorplan)
        .options(joinedload(Floorplan.serverroom_links))
        .filter(Floorplan.id == floorplan_id)
        .first()
    )
    if not floorplan:
        raise HTTPException(status_code=404, detail="Floorplan not found")

    old_values = {"name": floorplan.name, "width": floorplan.width, "height": floorplan.height}
    for key, value in payload.model_dump().items():
        setattr(floorplan, key, value)

    write_audit(
        db,
        actor=user,
        action="update",
        entity_type="floorplan",
        entity_id=str(floorplan.id),
        old_values=old_values,
        new_values=payload.model_dump(),
    )
    db.commit()
    db.refresh(floorplan)
    out = FloorplanOut.model_validate(floorplan)
    out.serverroom_id = map_floorplan_serverroom_id(floorplan)
    return out


@app.post("/api/racks", response_model=RackOut)
def create_rack(
    payload: RackCreate,
    db: Session = Depends(get_db),
    user: str = Depends(get_current_user),
) -> RackOut:
    floorplan = db.query(Floorplan).filter(Floorplan.id == payload.floorplan_id).first()
    if not floorplan:
        raise HTTPException(status_code=404, detail="Floorplan not found")

    rack = Rack(**payload.model_dump())
    db.add(rack)
    db.flush()
    write_audit(
        db,
        actor=user,
        action="create",
        entity_type="rack",
        entity_id=str(rack.id),
        new_values=payload.model_dump(),
    )
    db.commit()
    db.refresh(rack)
    return rack


@app.put("/api/racks/{rack_id}", response_model=RackOut)
def update_rack(
    rack_id: int,
    payload: RackUpdate,
    db: Session = Depends(get_db),
    user: str = Depends(get_current_user),
) -> RackOut:
    rack = db.query(Rack).filter(Rack.id == rack_id).first()
    if not rack:
        raise HTTPException(status_code=404, detail="Rack not found")

    old_values = {
        "name": rack.name,
        "x": rack.x,
        "y": rack.y,
        "width": rack.width,
        "height": rack.height,
        "units": rack.units,
        "orientation": rack.orientation,
    }
    for key, value in payload.model_dump().items():
        setattr(rack, key, value)

    write_audit(
        db,
        actor=user,
        action="update",
        entity_type="rack",
        entity_id=str(rack.id),
        old_values=old_values,
        new_values=payload.model_dump(),
    )
    db.commit()
    db.refresh(rack)
    return rack


@app.delete("/api/racks/{rack_id}")
def delete_rack(
    rack_id: int,
    db: Session = Depends(get_db),
    user: str = Depends(get_current_user),
) -> dict[str, bool]:
    rack = db.query(Rack).filter(Rack.id == rack_id).first()
    if not rack:
        raise HTTPException(status_code=404, detail="Rack not found")

    old_values = {
        "name": rack.name,
        "x": rack.x,
        "y": rack.y,
        "width": rack.width,
        "height": rack.height,
        "units": rack.units,
        "orientation": rack.orientation,
    }
    db.delete(rack)
    write_audit(
        db,
        actor=user,
        action="delete",
        entity_type="rack",
        entity_id=str(rack_id),
        old_values=old_values,
    )
    db.commit()
    return {"ok": True}


@app.post("/api/devices", response_model=DeviceOut)
def create_device(
    payload: DeviceCreate,
    db: Session = Depends(get_db),
    user: str = Depends(get_current_user),
) -> DeviceOut:
    ensure_serial_not_archived(db, payload.serial_number)
    validate_device_placement(db, payload.rack_id, payload.u_position, payload.u_height, payload.mount_side)

    payload_dict = payload.model_dump()
    model_id = payload_dict.pop("device_model_id", None)
    if model_id is not None:
        template = db.query(DeviceModel).filter(DeviceModel.id == model_id).first()
        if not template:
            raise HTTPException(status_code=404, detail="Device model not found")
        payload_dict["name"] = payload_dict["name"] or template.name
        payload_dict["device_type"] = template.device_type
        payload_dict["u_height"] = template.u_height
        payload_dict["model"] = template.model_code
        payload_dict["vendor"] = template.vendor
        if payload_dict.get("mount_side") == "front" and template.device_type == "switch":
            payload_dict["mount_side"] = "back"
        properties = payload_dict.get("properties", {})
        properties["image_url"] = template.image_url
        payload_dict["properties"] = properties

    device = Device(**payload_dict)
    db.add(device)
    db.flush()
    write_audit(
        db,
        actor=user,
        action="create",
        entity_type="device",
        entity_id=str(device.id),
        new_values=payload.model_dump(),
    )
    db.commit()
    db.refresh(device)
    return device


@app.put("/api/devices/{device_id}", response_model=DeviceOut)
def update_device(
    device_id: int,
    payload: DeviceUpdate,
    db: Session = Depends(get_db),
    user: str = Depends(get_current_user),
) -> DeviceOut:
    device = db.query(Device).filter(Device.id == device_id).first()
    if not device:
        raise HTTPException(status_code=404, detail="Device not found")

    ensure_serial_not_archived(db, payload.serial_number)

    validate_device_placement(
        db,
        device.rack_id,
        payload.u_position,
        payload.u_height,
        payload.mount_side,
        exclude_device_id=device.id,
    )

    old_values = {
        "name": device.name,
        "device_type": device.device_type,
        "u_position": device.u_position,
        "u_height": device.u_height,
        "mount_side": device.mount_side,
        "serial_number": device.serial_number,
        "management_ip": device.management_ip,
        "model": device.model,
        "vendor": device.vendor,
        "properties": device.properties,
    }
    for key, value in payload.model_dump().items():
        setattr(device, key, value)

    write_audit(
        db,
        actor=user,
        action="update",
        entity_type="device",
        entity_id=str(device.id),
        old_values=old_values,
        new_values=payload.model_dump(),
    )
    db.commit()
    db.refresh(device)
    return device


@app.delete("/api/devices/{device_id}")
def delete_device(
    device_id: int,
    db: Session = Depends(get_db),
    user: str = Depends(get_current_user),
) -> dict[str, bool]:
    device = db.query(Device).filter(Device.id == device_id).first()
    if not device:
        raise HTTPException(status_code=404, detail="Device not found")

    old_values = {
        "name": device.name,
        "device_type": device.device_type,
        "u_position": device.u_position,
        "u_height": device.u_height,
        "mount_side": device.mount_side,
        "serial_number": device.serial_number,
        "management_ip": device.management_ip,
        "model": device.model,
        "vendor": device.vendor,
        "properties": device.properties,
    }
    db.delete(device)
    write_audit(
        db,
        actor=user,
        action="delete",
        entity_type="device",
        entity_id=str(device_id),
        old_values=old_values,
    )
    db.commit()
    return {"ok": True}


@app.post("/api/devices/{device_id}/unrack")
def unrack_device(
    device_id: int,
    db: Session = Depends(get_db),
    user: str = Depends(get_current_user),
) -> dict[str, object]:
    device = db.query(Device).filter(Device.id == device_id).first()
    if not device:
        raise HTTPException(status_code=404, detail="Device not found")

    old_values = {
        "name": device.name,
        "device_type": device.device_type,
        "u_position": device.u_position,
        "u_height": device.u_height,
        "mount_side": device.mount_side,
        "serial_number": device.serial_number,
        "management_ip": device.management_ip,
        "model": device.model,
        "vendor": device.vendor,
        "properties": device.properties,
    }

    inventory_serial = upsert_inventory_from_device(db, device, archived=False)
    db.delete(device)
    write_audit(
        db,
        actor=user,
        action="unrack",
        entity_type="device",
        entity_id=str(device_id),
        old_values=old_values,
        new_values={"inventory_serial": inventory_serial},
    )
    db.commit()
    return {"ok": True, "inventory_serial": inventory_serial}


@app.post("/api/devices/{device_id}/archive")
def archive_device(
    device_id: int,
    db: Session = Depends(get_db),
    user: str = Depends(get_current_user),
) -> dict[str, object]:
    device = db.query(Device).filter(Device.id == device_id).first()
    if not device:
        raise HTTPException(status_code=404, detail="Device not found")

    old_values = {
        "name": device.name,
        "device_type": device.device_type,
        "u_position": device.u_position,
        "u_height": device.u_height,
        "mount_side": device.mount_side,
        "serial_number": device.serial_number,
        "management_ip": device.management_ip,
        "model": device.model,
        "vendor": device.vendor,
        "properties": device.properties,
    }

    inventory_serial = upsert_inventory_from_device(db, device, archived=True)

    db.delete(device)
    write_audit(
        db,
        actor=user,
        action="archive",
        entity_type="device",
        entity_id=str(device_id),
        old_values=old_values,
        new_values={"inventory_serial": inventory_serial, "archived": True},
    )
    db.commit()
    return {"ok": True, "inventory_serial": inventory_serial, "archived": True}


@app.post("/api/import/inventory-csv")
def import_inventory_csv(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: str = Depends(get_current_user),
) -> dict[str, object]:
    rows = parse_csv_upload(file)
    created = 0
    updated = 0
    errors: list[str] = []

    for idx, row in enumerate(rows, start=2):
        serial = csv_value(row, ["serial_number", "serialnumber", "serial", "sn"])
        name = csv_value(row, ["name", "device_name", "hostname"])

        if not serial:
            errors.append(f"line {idx}: missing serial number")
            continue
        if not name:
            errors.append(f"line {idx}: missing device name")
            continue

        try:
            u_height = csv_int_value(row, ["u_height", "uheight", "height_u", "u"], default=1)
        except ValueError as err:
            errors.append(f"line {idx}: {err}")
            continue

        if u_height < 1:
            errors.append(f"line {idx}: u_height must be >= 1")
            continue

        payload = {
            "name": name,
            "model": csv_value(row, ["model", "model_code"], default=""),
            "vendor": csv_value(row, ["vendor", "manufacturer"], default=""),
            "device_type": csv_value(row, ["device_type", "type"], default="server") or "server",
            "u_height": u_height,
            "mount_side": csv_value(row, ["mount_side", "side", "position"], default="").strip().lower(),
            "management_ip": csv_value(row, ["management_ip", "mgmt_ip", "idrac", "ilo"], default=""),
            "properties": {
                "hostname": csv_value(row, ["hostname"], default=""),
                "host_ip": csv_value(row, ["host_ip", "ip"], default=""),
                "ssh_endpoint": csv_value(row, ["ssh_endpoint", "ssh"], default=""),
                "notes": csv_value(row, ["notes", "comment"], default=""),
                "image_url": csv_value(row, ["image_url", "image"], default=""),
            },
        }

        if payload["mount_side"] not in {"front", "back"}:
            payload["mount_side"] = "back" if payload["device_type"] == "switch" else "front"

        existing = db.query(InventoryDevice).filter(InventoryDevice.serial_number == serial).first()
        if existing:
            existing.name = payload["name"]
            existing.model = payload["model"] or None
            existing.vendor = payload["vendor"] or None
            existing.device_type = payload["device_type"]
            existing.u_height = payload["u_height"]
            existing.mount_side = payload["mount_side"]
            existing.management_ip = payload["management_ip"] or None
            existing.properties = payload["properties"]
            updated += 1
        else:
            db.add(
                InventoryDevice(
                    serial_number=serial,
                    name=payload["name"],
                    model=payload["model"] or None,
                    vendor=payload["vendor"] or None,
                    device_type=payload["device_type"],
                    u_height=payload["u_height"],
                    mount_side=payload["mount_side"],
                    management_ip=payload["management_ip"] or None,
                    properties=payload["properties"],
                )
            )
            created += 1

    write_audit(
        db,
        actor=user,
        action="import",
        entity_type="inventory_csv",
        entity_id="bulk",
        new_values={"created": created, "updated": updated, "errors": len(errors)},
    )
    db.commit()

    return {
        "created": created,
        "updated": updated,
        "errors": errors,
        "rows": len(rows),
    }


@app.post("/api/import/layout-csv")
def import_layout_csv(
    file: UploadFile = File(...),
    clear_existing: bool = True,
    db: Session = Depends(get_db),
    user: str = Depends(get_current_user),
) -> dict[str, object]:
    rows = parse_csv_upload(file)
    created = 0
    updated = 0
    errors: list[str] = []
    cleared_rack_ids: set[int] = set()

    for idx, row in enumerate(rows, start=2):
        room_name = csv_value(row, ["serverroom", "server_room", "room"])
        floor_name = csv_value(row, ["floor", "floorplan", "floor_name"])
        rack_name = csv_value(row, ["rackname", "rack", "rack_name"])
        serial = csv_value(row, ["serial_number", "serialnumber", "serial", "sn"])

        if not room_name or not floor_name or not rack_name or not serial:
            errors.append(
                f"line {idx}: required fields are serverroom, floor, rackname, serialnumber"
            )
            continue

        try:
            u_position = csv_int_value(row, ["u_position", "uposition", "u", "position"])
        except ValueError as err:
            errors.append(f"line {idx}: {err}")
            continue

        if u_position < 1:
            errors.append(f"line {idx}: u_position must be >= 1")
            continue

        room = find_serverroom_by_name(db, room_name)
        if room is None:
            errors.append(f"line {idx}: serverroom '{room_name}' not found")
            continue

        floorplan = find_floorplan_by_room_and_name(db, room.id, floor_name)
        if floorplan is None:
            errors.append(f"line {idx}: floor '{floor_name}' not found in serverroom '{room_name}'")
            continue

        rack = find_rack_by_floorplan_and_name(db, floorplan.id, rack_name)
        if rack is None:
            errors.append(f"line {idx}: rack '{rack_name}' not found on floor '{floor_name}'")
            continue

        inventory = db.query(InventoryDevice).filter(InventoryDevice.serial_number == serial).first()
        if inventory is None:
            errors.append(f"line {idx}: serial '{serial}' not found in inventory import")
            continue
        if inventory.archived == 1:
            errors.append(f"line {idx}: serial '{serial}' is archived and cannot be assigned")
            continue

        if clear_existing and rack.id not in cleared_rack_ids:
            db.query(Device).filter(Device.rack_id == rack.id).delete(synchronize_session=False)
            cleared_rack_ids.add(rack.id)

        existing_device = db.query(Device).filter(Device.serial_number == serial).first()
        try:
            if existing_device:
                validate_device_placement(
                    db,
                    rack.id,
                    u_position,
                    inventory.u_height,
                    inventory.mount_side or "front",
                    exclude_device_id=existing_device.id,
                )
                existing_device.rack_id = rack.id
                existing_device.name = inventory.name
                existing_device.device_type = inventory.device_type
                existing_device.u_position = u_position
                existing_device.u_height = inventory.u_height
                existing_device.mount_side = inventory.mount_side or "front"
                existing_device.serial_number = inventory.serial_number
                existing_device.management_ip = inventory.management_ip
                existing_device.model = inventory.model
                existing_device.vendor = inventory.vendor
                existing_device.properties = inventory.properties or {}
                updated += 1
            else:
                validate_device_placement(
                    db,
                    rack.id,
                    u_position,
                    inventory.u_height,
                    inventory.mount_side or "front",
                )
                db.add(
                    Device(
                        rack_id=rack.id,
                        name=inventory.name,
                        device_type=inventory.device_type,
                        u_position=u_position,
                        u_height=inventory.u_height,
                        mount_side=inventory.mount_side or "front",
                        serial_number=inventory.serial_number,
                        management_ip=inventory.management_ip,
                        model=inventory.model,
                        vendor=inventory.vendor,
                        properties=inventory.properties or {},
                    )
                )
                created += 1
        except HTTPException as err:
            errors.append(f"line {idx}: {err.detail}")

    write_audit(
        db,
        actor=user,
        action="import",
        entity_type="layout_csv",
        entity_id="bulk",
        new_values={"created": created, "updated": updated, "errors": len(errors)},
    )
    db.commit()

    return {
        "created": created,
        "updated": updated,
        "errors": errors,
        "rows": len(rows),
    }


@app.get("/api/audit", response_model=list[AuditOut])
def list_audit(
    limit: int = 100,
    db: Session = Depends(get_db),
    _: str = Depends(get_current_user),
) -> list[AuditOut]:
    logs = db.query(AuditLog).order_by(AuditLog.created_at.desc()).limit(limit).all()
    return logs


@app.post("/api/admin/clear-all-data")
def clear_all_data(
    payload: PasswordConfirmRequest,
    db: Session = Depends(get_db),
    user: str = Depends(require_admin),
) -> dict[str, bool]:
    # Re-authenticate the admin with provided password to confirm intent
    admin_user = authenticate_local_user(db, user, payload.password)
    if not admin_user or admin_user.role != "admin":
        raise HTTPException(status_code=403, detail="Invalid admin password")

    # Delete all data except local users in dependency order
    db.query(AuditLog).delete()
    db.query(Device).delete()
    db.query(Rack).delete()
    db.query(ServerRoomFloorplan).delete()
    db.query(Floorplan).delete()
    db.query(ServerRoom).delete()
    db.query(InventoryDevice).delete()
    db.query(DeviceModel).delete()

    write_audit(
        db,
        actor=user,
        action="clear-all",
        entity_type="system",
        entity_id="*",
        old_values={"note": "All data cleared by admin"},
    )
    db.commit()
    return {"ok": True}
